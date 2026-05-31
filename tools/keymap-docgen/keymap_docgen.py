"""
keymap_docgen.py

ZMK keymap (.keymap) のレイヤー（指定がなければ全レイヤー）の全キー割り当てを以下 2 形式で
出力するドキュメンテーション生成ツール：

  1. Excel ファイル (.xlsx)
       - "動作" シートと "経路" シートを生成し、各シートに全レイヤーの表を縦に並べる
  2. 自己完結型 HTML ファイル (.html)
       - 同じ内容を HTML の表で出力（「Row N」見出し行を背景色でハイライト）
       - 複数レイヤー指定時は「動作」セクションに全レイヤーを並べた後、
         「経路」セクションに全レイヤーを再度並べる構成

それぞれの表は、キーボード物理行ごとに以下の構造を持つ：
  - 左端 1 列: 「操作」 = タップ / ホールド / ダブルタップ / Shift+ / Ctrl+
  - 右側の列: その物理行のキーを物理配列順に並べたもの

mod-morph (LSHIFT/RSHIFT, LCTL/RCTL), tap-dance, layer-tap, momentary-layer
など標準的な ZMK behavior を解析し、再帰的に動作を解決する。

このスクリプトはキーボード固有のデータを一切持たない。各キーの物理位置と表示
ラベルは -l で渡す物理配列 JSON（各エントリの x / y と任意の label）から取得する
ため、同じスクリプトを複数のキーボードでそのまま再利用できる。

Usage:
    python keymap_docgen.py <keymap_file> [<layer_name> ...] [-l layout.json] [-o output.xlsx]

    レイヤー名を省略すると、keymap 内の全レイヤーを定義順で出力する。
    -l を省略すると keymap と同じベース名の .json を探し、見つからなければ
    keymap 順の 1 行レイアウトにフォールバックする。

出力先：
    -o で指定した .xlsx と同じディレクトリ／同じベース名で .html も生成される
    （-o を省略するとレイヤー単独時は <layer>_keymap.xlsx、それ以外は keymap.xlsx）

Examples:
    # 全レイヤーを定義順で出力（レイヤー名の指定を省略）
    python keymap_docgen.py <keymap_file> -l <layout.json> -o KEYMAP.xlsx
    # 特定のレイヤーだけ出力
    python keymap_docgen.py <keymap_file> <LAYER_NAME> -l <layout.json> -o KEYMAP.xlsx
    python keymap_docgen.py <keymap_file> <LAYER_A> <LAYER_B> <LAYER_C> -l <layout.json> -o KEYMAP.xlsx
"""

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:  # Excel 出力は任意。未インストールでも Markdown は生成できる
    HAVE_OPENPYXL = False


# ============================================================================
# Parsing
# ============================================================================

def strip_comments(text: str) -> str:
    text = re.sub(r'//[^\n]*', '', text)
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    return text


# ============================================================================
# #define expansion + readable layer-name display
# ============================================================================

# Index -> layer node name (e.g. 2 -> 'VIM_NORMAL_1'). Populated in main() from
# the keymap's definition order. Used to render &lt/&mo/&to routes with the
# formal layer name instead of a bare number.
LAYER_NAMES_BY_INDEX: dict[int, str] = {}


def parse_defines(text: str) -> dict[str, int]:
    """Collect integer `#define NAME N` macros (layer aliases) from the keymap."""
    return {m.group(1): int(m.group(2))
            for m in re.finditer(r'(?m)^\s*#define\s+(\w+)\s+(\d+)\s*$', text)}


def expand_defines(text: str, defines: dict[str, int]) -> str:
    """Drop `#define` lines and substitute each defined name (whole word) with its
    numeric value, so the rest of the tool can keep assuming numeric layer ids."""
    text = re.sub(r'(?m)^[ \t]*#define[ \t]+\w+[ \t]+\d+[ \t]*\n?', '', text)
    for name in sorted(defines, key=len, reverse=True):
        text = re.sub(rf'\b{re.escape(name)}\b', str(defines[name]), text)
    return text


def layer_display(n) -> str:
    """Render a layer index as its formal layer name (fallback: L<n>)."""
    try:
        i = int(n)
    except (TypeError, ValueError):
        return str(n)
    return LAYER_NAMES_BY_INDEX.get(i, f'L{n}')


def format_binding_for_display(b: str) -> str:
    """Rewrite numeric layer ids in a raw binding to formal layer names for display
    (e.g. '&lt 2 SPACE' -> '&lt VIM_NORMAL_1 SPACE', '&mo 7' -> '&mo BLUETOOTH')."""
    return re.sub(r'(&(?:lt|mo|to))\s+(\d+)',
                  lambda m: f'{m.group(1)} {layer_display(m.group(2))}', b)


def find_balanced_block(content: str, start_pos: int) -> tuple[int, int] | None:
    """Find the position of the matching close brace for an open brace at start_pos."""
    if start_pos >= len(content) or content[start_pos] != '{':
        return None
    depth = 1
    i = start_pos + 1
    while i < len(content):
        if content[i] == '{':
            depth += 1
        elif content[i] == '}':
            depth -= 1
            if depth == 0:
                return (start_pos + 1, i)
        i += 1
    return None


def extract_named_block(content: str, name: str) -> str | None:
    """Extract the content of a top-level named block like 'macros { ... }'."""
    # Find 'name {'
    for m in re.finditer(r'\b' + re.escape(name) + r'\s*\{', content):
        brace_pos = m.end() - 1
        range_ = find_balanced_block(content, brace_pos)
        if range_:
            return content[range_[0]:range_[1]]
    return None


def parse_definitions(block_content: str) -> dict:
    """
    Parse 'name: label { ... };' definitions inside a block.
    Returns dict: {name: body_string}
    """
    defs = {}
    i = 0
    while i < len(block_content):
        m = re.search(r'(\w+)\s*:\s*\w+\s*\{', block_content[i:])
        if not m:
            break
        name = m.group(1)
        brace_pos = i + m.end() - 1
        range_ = find_balanced_block(block_content, brace_pos)
        if not range_:
            break
        body = block_content[range_[0]:range_[1]]
        defs[name] = body
        i = range_[1] + 1
    return defs


def parse_bindings_list(body: str) -> list[str]:
    """Extract list of bindings from a 'bindings = <...>, <...>;' style body."""
    m = re.search(r'bindings\s*=\s*([^;]+);', body, re.DOTALL)
    if not m:
        return []
    text = m.group(1)
    return [s.strip() for s in re.findall(r'<([^>]+)>', text)]


def parse_compatible(body: str) -> str | None:
    m = re.search(r'compatible\s*=\s*"([^"]+)"', body)
    return m.group(1) if m else None


def parse_mods(body: str) -> str | None:
    m = re.search(r'mods\s*=\s*<\(?([^>)]+)\)?>', body)
    return m.group(1).strip() if m else None


def parse_macros(content: str) -> dict:
    block = extract_named_block(content, 'macros')
    if not block:
        return {}
    defs = parse_definitions(block)
    return {name: {'bindings': parse_bindings_list(body)} for name, body in defs.items()}


def parse_behaviors(content: str) -> dict:
    block = extract_named_block(content, 'behaviors')
    if not block:
        return {}
    defs = parse_definitions(block)
    out = {}
    for name, body in defs.items():
        out[name] = {
            'compatible': parse_compatible(body),
            'bindings': parse_bindings_list(body),
            'mods': parse_mods(body),
        }
    return out


def parse_layer(content: str, layer_name: str) -> str | None:
    """Extract the raw bindings text of a named layer."""
    keymap_block = extract_named_block(content, 'keymap')
    if not keymap_block:
        return None
    defs = {}
    # layers are 'NAME { ... }' (no label after colon)
    i = 0
    while i < len(keymap_block):
        m = re.search(r'(\w+)\s*\{', keymap_block[i:])
        if not m:
            break
        name = m.group(1)
        brace_pos = i + m.end() - 1
        range_ = find_balanced_block(keymap_block, brace_pos)
        if not range_:
            break
        body = keymap_block[range_[0]:range_[1]]
        if name == layer_name:
            m2 = re.search(r'bindings\s*=\s*<(.+?)>', body, re.DOTALL)
            if m2:
                return m2.group(1)
        i = range_[1] + 1
    return None


def parse_all_layer_names(content: str) -> list[str]:
    """Return all layer names defined in the keymap block, in definition order."""
    keymap_block = extract_named_block(content, 'keymap')
    if not keymap_block:
        return []
    names = []
    i = 0
    while i < len(keymap_block):
        m = re.search(r'(\w+)\s*\{', keymap_block[i:])
        if not m:
            break
        brace_pos = i + m.end() - 1
        range_ = find_balanced_block(keymap_block, brace_pos)
        if not range_:
            break
        body = keymap_block[range_[0]:range_[1]]
        # Only treat blocks that actually have bindings as layers.
        if re.search(r'bindings\s*=\s*<', body):
            names.append(m.group(1))
        i = range_[1] + 1
    return names


def split_layer_bindings(text: str) -> list[str]:
    """Split a layer bindings string into individual binding strings (each starts with '&')."""
    text = re.sub(r'\s+', ' ', text).strip()
    out = []
    i = 0
    n = len(text)
    while i < n:
        if text[i] == '&':
            j = i + 1
            while j < n and text[j] != '&':
                j += 1
            out.append(text[i:j].strip())
            i = j
        else:
            i += 1
    return out


# ============================================================================
# Keycode formatting
# ============================================================================

KEYCODE_LABELS = {
    'LEFT': '←', 'RIGHT': '→', 'UP_ARROW': '↑', 'UP': '↑', 'DOWN': '↓',
    'HOME': 'HOME', 'END': 'END', 'ENTER': 'ENTER', 'DELETE': 'DEL',
    'BACKSPACE': 'BS', 'TAB': 'TAB', 'SPACE': 'SPACE', 'ESCAPE': 'ESC',
    'PAGE_UP': 'PgUp', 'PAGE_DOWN': 'PgDn',
    'LCTRL': 'LCtrl', 'RCTRL': 'RCtrl',
    'LSHIFT': 'LShift', 'RSHIFT': 'RShift',
    'LEFT_SHIFT': 'LShift', 'RIGHT_SHIFT': 'RShift',
    'LEFT_CONTROL': 'LCtrl', 'RIGHT_CONTROL': 'RCtrl',
    'LEFT_ALT': 'LAlt', 'RIGHT_ALT': 'RAlt',
    'LEFT_WIN': 'LWin', 'RIGHT_WIN': 'RWin',
    'GREATER_THAN': '>', 'LESS_THAN': '<',
    'COMMA': ',', 'PERIOD': '.', 'SEMICOLON': ';', 'SLASH': '/',
    'SINGLE_QUOTE': "'", 'GRAVE': '`', 'BACKSLASH': '\\', 'EQUAL': '=',
    'MINUS': '-', 'LEFT_BRACKET': '[', 'RIGHT_BRACKET': ']',
    'N0': '0', 'N1': '1', 'N2': '2', 'N3': '3', 'N4': '4',
    'N5': '5', 'N6': '6', 'N7': '7', 'N8': '8', 'N9': '9',
}

MOD_PREFIX = {
    'LC': '⌃', 'RC': '⌃',
    'LS': '⇧', 'RS': '⇧',
    'LA': '⌥', 'RA': '⌥',
    'LG': '⌘', 'RG': '⌘',
}


def format_keycode(kc: str) -> str:
    kc = kc.strip()
    m = re.match(r'(LC|LS|LA|LG|RC|RS|RA|RG)\((.+)\)$', kc)
    if m:
        return f"{MOD_PREFIX[m.group(1)]}{format_keycode(m.group(2))}"
    return KEYCODE_LABELS.get(kc, kc)


# ============================================================================
# Resolution: binding × operation -> (action_description, path)
# ============================================================================

OPS = ('タップ', 'ホールド', 'ダブルタップ', 'Shift+', 'Ctrl+')


def resolve(binding: str, behaviors: dict, macros: dict, op: str, depth: int = 0) -> tuple[str, str]:
    """Resolve a binding string for the given operation."""
    if depth > 10:
        return ('再帰深度超過', binding)
    b = binding.strip()

    if b == '&none':
        return ('何もしない', '&none')

    if b == '&trans':
        return ('フォールスルー', '&trans')

    # &kp X
    m = re.match(r'&kp\s+(.+)$', b)
    if m:
        kc = m.group(1).strip()
        label = format_keycode(kc)
        path = f'&kp {kc}'
        if op == 'タップ':
            return (label, path)
        if op == 'ホールド':
            return (label, path)
        if op == 'ダブルタップ':
            return (f'{label}×2', path)
        if op == 'Shift+':
            return (f'⇧{label}', path)
        if op == 'Ctrl+':
            return (f'⌃{label}', path)

    # &mt MOD KEY
    m = re.match(r'&mt\s+(\S+)\s+(.+)$', b)
    if m:
        mod, key = m.group(1).strip(), m.group(2).strip()
        key_label = format_keycode(key)
        mod_label = format_keycode(mod)
        path = f'&mt {mod} {key}'
        if op == 'タップ':
            return (key_label, path)
        if op == 'ホールド':
            return (mod_label, path)
        if op == 'ダブルタップ':
            return (f'{key_label}×2', path)
        if op == 'Shift+':
            return (f'⇧{key_label}', path)
        if op == 'Ctrl+':
            return (f'⌃{key_label}', path)

    # &lt LAYER KEY
    m = re.match(r'&lt\s+(\d+)\s+(.+)$', b)
    if m:
        layer, key = m.group(1), m.group(2).strip()
        key_label = format_keycode(key)
        path = f'&lt {layer_display(layer)} {key}'
        if op == 'タップ':
            return (key_label, path)
        if op == 'ホールド':
            return (f'L{layer}', path)
        if op == 'ダブルタップ':
            return (f'{key_label}×2', path)
        if op == 'Shift+':
            return (f'⇧{key_label}', path)
        if op == 'Ctrl+':
            return (f'⌃{key_label}', path)

    # &mo X
    m = re.match(r'&mo\s+(\d+)$', b)
    if m:
        layer = m.group(1)
        return (f'L{layer}', f'&mo {layer_display(layer)}')

    # &to X
    m = re.match(r'&to\s+(\d+)$', b)
    if m:
        layer = m.group(1)
        return (f'⇒{layer_display(layer)}', f'&to {layer_display(layer)}')

    # Custom behavior / macro reference like &mm_vim_g, &td_vim_d, &macro_vim_dd
    if b.startswith('&'):
        name = b[1:].split()[0]
        if name in behaviors:
            return resolve_behavior(name, behaviors, macros, op, depth + 1)
        if name in macros:
            return resolve_macro(name, behaviors, macros, op, depth + 1)

    return (f'未対応: {b}', b)


def resolve_behavior(name: str, behaviors: dict, macros: dict, op: str, depth: int) -> tuple[str, str]:
    beh = behaviors[name]
    compat = beh['compatible']
    bindings = beh['bindings']
    mods = beh['mods'] or ''

    if compat == 'zmk,behavior-mod-morph':
        is_shift = 'LSFT' in mods or 'RSFT' in mods
        is_ctrl = 'LCTL' in mods or 'RCTL' in mods

        if op in ('タップ', 'ダブルタップ', 'ホールド'):
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, op, depth)
            return (sub_a, f'{name}[0] ▸ {sub_p}')

        if op == 'Shift+':
            if is_shift:
                sub_a, sub_p = resolve(bindings[1], behaviors, macros, 'タップ', depth)
                return (sub_a, f'{name}[1] ▸ {sub_p}')
            else:
                sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Shift+', depth)
                return (sub_a, f'{name}[0] ▸ {sub_p}')

        if op == 'Ctrl+':
            if is_ctrl:
                sub_a, sub_p = resolve(bindings[1], behaviors, macros, 'タップ', depth)
                return (sub_a, f'{name}[1] ▸ {sub_p}')
            else:
                sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Ctrl+', depth)
                return (sub_a, f'{name}[0] ▸ {sub_p}')

    if compat == 'zmk,behavior-tap-dance':
        if op == 'タップ':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'タップ', depth)
            return (sub_a, f'{name}[0] ▸ {sub_p}')
        if op == 'ホールド':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'ホールド', depth)
            return (sub_a, f'{name}[0] ▸ {sub_p}')
        if op == 'ダブルタップ':
            sub_a, sub_p = resolve(bindings[1], behaviors, macros, 'タップ', depth)
            return (sub_a, f'{name}[1] ▸ {sub_p}')
        if op == 'Shift+':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Shift+', depth)
            return (sub_a, f'{name}[0] ▸ {sub_p}')
        if op == 'Ctrl+':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Ctrl+', depth)
            return (sub_a, f'{name}[0] ▸ {sub_p}')

    return (f'未対応 behavior: {compat}', name)


def resolve_macro(name: str, behaviors: dict, macros: dict, op: str, depth: int) -> tuple[str, str]:
    summary = summarize_macro(macros[name]['bindings'])
    if op == 'ダブルタップ':
        return (f'{summary}×2', name)
    if op == 'Shift+':
        return (f'⇧{summary}', name)
    if op == 'Ctrl+':
        return (f'⌃{summary}', name)
    return (summary, name)


def summarize_macro(bindings: list[str]) -> str:
    parts = []
    for b in bindings:
        b = b.strip()
        if b.startswith('&macro_wait_time'):
            continue
        if b.startswith('&kp '):
            for kc in re.findall(r'&kp\s+([A-Z0-9_()]+)', b):
                parts.append(format_keycode(kc))
        elif b.startswith('&macro_release'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(f'{format_keycode(inner.group(1))} 解放')
        elif b.startswith('&macro_press'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(f'{format_keycode(inner.group(1))} 押下')
        elif b.startswith('&macro_tap'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(format_keycode(inner.group(1)))
        elif b.startswith('&to '):
            parts.append(f'⇒{layer_display(b[4:].strip())}')
        else:
            parts.append(b)
    return ' ▸ '.join(parts)


# ============================================================================
# Visual label (per physical key index, not per column position)
# ============================================================================

# Per-binding-index display label = that physical key's identity on the DEFAULT
# layer (e.g. index 0 -> 'Q'). Populated at runtime in main() from the physical-
# layout JSON's optional per-key "label" field, so this script carries no
# keyboard-specific data. Keyed by index (not by column slot) so reordering
# columns into physical order can never desync a label from its binding.
KEY_LABELS: dict[int, str] = {}


def get_label(idx: int) -> str:
    return KEY_LABELS.get(idx, f'pos {idx}')


# ============================================================================
# Excel output
# ============================================================================

def get_row_layout(total: int) -> list[tuple[int, int]]:
    """Fallback row layout when no physical-layout JSON is supplied.

    Without per-key (x, y) coordinates the real row split is unknown, so place
    every binding in a single keymap-order row. Pass a physical-layout JSON
    (see load_physical_layout) to get an accurate multi-row grid."""
    return [(0, total)]  # fallback: linear (single keymap-order row)


# ============================================================================
# Physical layout grid (drives row/column arrangement from real coordinates)
# ============================================================================

GAP = 'GAP'  # sentinel marking a blank split-gap display column


def load_physical_layout(path: Path):
    """Parse a ZMK physical-layout JSON into ordered per-key data.

    The JSON's `layout` array is in the same order as the keymap bindings, so
    entry i describes binding i. Each entry carries `x`/`y` (physical position)
    and may carry an optional `label` (that key's DEFAULT-layer identity, e.g.
    "Q"), which is how the tool stays free of keyboard-specific data.

    Returns a (coords, labels) pair:
      - coords: list of (x, y) floats, or None when the file is missing or
        cannot be parsed (caller falls back to keymap-order rendering).
      - labels: {binding_index: label} for every entry that provides a label.
    """
    try:
        data = json.loads(Path(path).read_text(encoding='utf-8'))
    except (OSError, ValueError):
        return None, {}
    layouts = data.get('layouts') if isinstance(data, dict) else None
    if not layouts:
        return None, {}
    layout = layouts.get('default_layout') or next(iter(layouts.values()))
    entries = layout.get('layout') if isinstance(layout, dict) else None
    if not entries:
        return None, {}
    coords: list[tuple[float, float]] = []
    labels: dict[int, str] = {}
    for i, e in enumerate(entries):
        try:
            coords.append((float(e['x']), float(e['y'])))
        except (KeyError, TypeError, ValueError):
            return None, {}
        if isinstance(e, dict) and e.get('label') is not None:
            labels[i] = str(e['label'])
    return (coords or None), labels


def build_grid(coords: list[tuple[float, float]]):
    """Build physical rows + display columns from per-index (x, y) coordinates.

    Rows are grouped by distinct y (top to bottom); columns are the distinct x
    values (left to right) with a single blank GAP column inserted at the widest
    horizontal gap (the split between the two halves). Returns
    (rows, display_cols) where:
      - display_cols: left-to-right list of x-values (float) and GAP sentinels.
      - rows: list of {'y': y, 'cells': [idx | None per display column]} where a
        None cell is an empty position, the split gap, or a missing key.
    """
    if not coords:
        return None, None
    xs = sorted({x for x, _ in coords})
    ys = sorted({y for _, y in coords})

    # Insert a gap column at the largest consecutive x-gap, if one clearly exists.
    diffs = [xs[i + 1] - xs[i] for i in range(len(xs) - 1)]
    if diffs and max(diffs) > min(diffs):
        gap_after = diffs.index(max(diffs))
    else:
        gap_after = -1

    display_cols: list = []
    for i, x in enumerate(xs):
        display_cols.append(x)
        if i == gap_after:
            display_cols.append(GAP)

    pos = {(x, y): idx for idx, (x, y) in enumerate(coords)}
    rows = []
    for y in ys:
        cells = [None if col == GAP else pos.get((col, y)) for col in display_cols]
        rows.append({'y': y, 'cells': cells})
    return rows, display_cols


def legacy_grid(total: int):
    """Fallback grid (no physical coords): sequential keymap-order chunking."""
    layout = get_row_layout(total)
    max_w = max((count for _, count in layout), default=0)
    rows = []
    idx = 0
    for _, count in layout:
        cells = [idx + p for p in range(count)] + [None] * (max_w - count)
        rows.append({'y': None, 'cells': cells})
        idx += count
    return rows, list(range(max_w))


def _write_mode_sheet(ws, layers_data: list[tuple[str, list[str]]],
                      behaviors: dict, macros: dict, mode: str,
                      grid, display_cols, active_indices, is_single: bool) -> None:
    """Write one mode sheet ('動作' or '経路') laid out to match the HTML output:
    a single column header at the top of the sheet, then every layer stacked
    below it sharing those columns. Each layer starts with a highlighted
    layer-name row, followed by 'Row N' heading rows (carrying the key label
    + binding) and the operation rows that differ from the auto-derived tap.
    """
    title_font = Font(bold=True, size=14, name='Yu Gothic UI')
    layer_font = Font(bold=True, size=12, name='Yu Gothic UI')
    layer_fill = PatternFill('solid', start_color='E1ECF4', end_color='E1ECF4', fill_type='solid')
    header_font = Font(bold=True, size=10, name='Yu Gothic UI')
    header_fill = PatternFill('solid', start_color='F6F8FA', end_color='F6F8FA', fill_type='solid')
    row_font = Font(bold=True, size=10, name='Yu Gothic UI')
    row_fill = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    op_font = Font(bold=True, size=9, name='Yu Gothic UI')
    body_font = Font(size=9, name='Yu Gothic UI')

    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    mode_label = '動作' if mode == 'action' else '経路'
    n_cols = len(display_cols)

    # Sheet title
    if is_single:
        layer_name = layers_data[0][0]
        title = f'{layer_name} レイヤー - {mode_label}'
    else:
        title = f'キー割り当て一覧 - {mode_label}'
    c = ws.cell(1, 1, title)
    c.font = title_font

    # Pre-compute each layer's rows; capture the shared column header from
    # whichever layer first produced one (display_cols-driven, so identical
    # for every layer).
    shared_header: list[str] | None = None
    layer_blocks: list[tuple[str, list[dict]]] = []
    for layer_name, bindings in layers_data:
        header, rows = _build_layer_mode_table(bindings, behaviors, macros, mode,
                                               grid, display_cols, active_indices)
        if header is None:
            continue
        shared_header = header
        layer_blocks.append((layer_name, rows))

    if shared_header is None:
        return

    # Column widths: size each key column to its widest single line so cells never
    # auto-wrap (only the explicit ▸ / heading line breaks split a cell).
    ws.column_dimensions['A'].width = 22
    col_max = [0] * n_cols
    for _, rows in layer_blocks:
        for row in rows:
            if row['kind'] == 'heading':
                for j, key in enumerate(row['keys']):
                    if key is not None:
                        col_max[j] = max(col_max[j], _disp_width(key[0]),
                                         _disp_width(format_binding_for_display(key[1])))
            else:
                for j, value in enumerate(row['values']):
                    for line in _split_cell_lines(value):
                        col_max[j] = max(col_max[j], _disp_width(line))
    for j in range(n_cols):
        width = 3 if display_cols[j] == GAP else max(8, col_max[j] + 2)
        ws.column_dimensions[get_column_letter(2 + j)].width = width

    r = 3
    # Single column header row (操作 | 1 | 2 | ...) at the top of the sheet.
    for col, text in enumerate(shared_header, start=1):
        c = ws.cell(r, col, text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    r += 1

    for layer_name, rows in layer_blocks:
        if not is_single:
            # Layer separator row spanning the whole table width.
            c = ws.cell(r, 1, f'{layer_name} レイヤー')
            c.font = layer_font
            c.fill = layer_fill
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = border
            for j in range(n_cols):
                cc = ws.cell(r, 2 + j, '')
                cc.fill = layer_fill
                cc.border = border
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + n_cols)
            r += 1

        for row in rows:
            if row['kind'] == 'heading':
                c = ws.cell(r, 1, row['desc'])
                c.font = row_font
                c.fill = row_fill
                c.alignment = Alignment(horizontal='left', vertical='center')
                c.border = border
                for j, key in enumerate(row['keys']):
                    value = '' if key is None else f'{key[0]}\n{format_binding_for_display(key[1])}'
                    cc = ws.cell(r, 2 + j, value)
                    cc.font = body_font
                    cc.fill = row_fill
                    cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cc.border = border
                ws.row_dimensions[r].height = 30
            else:
                c = ws.cell(r, 1, row['op'])
                c.font = op_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border
                max_lines = 1
                for j, value in enumerate(row['values']):
                    lines = _split_cell_lines(value)
                    max_lines = max(max_lines, len(lines))
                    cc = ws.cell(r, 2 + j, '\n'.join(lines))
                    cc.font = body_font
                    # 複数行にわたるセルは左寄せにして読みやすくする。
                    horiz = 'left' if len(lines) > 1 else 'center'
                    cc.alignment = Alignment(horizontal=horiz, vertical='center', wrap_text=True)
                    cc.border = border
                # Give multi-line route cells enough height to show every step.
                if max_lines > 1:
                    ws.row_dimensions[r].height = 15 * max_lines
            r += 1


def write_excel(layers_data: list[tuple[str, list[str]]],
                behaviors: dict, macros: dict, output_path: Path,
                grid, display_cols) -> None:
    """Generate one Excel file matching the HTML/Markdown layout: two sheets
    '動作' / '経路', each stacking every layer's table vertically."""
    wb = Workbook()
    is_single = len(layers_data) == 1
    active_indices = None if is_single else _compute_active_indices(layers_data)
    first = True
    for _, mode in [('動作', 'action'), ('経路', 'path')]:
        sheet_name = '動作' if mode == 'action' else '経路'
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        _write_mode_sheet(ws, layers_data, behaviors, macros, mode,
                          grid, display_cols, active_indices, is_single)
    wb.save(output_path)


# ============================================================================
# Shared table data + HTML output
# ============================================================================

def _auto_forms(tap_value: str, op: str) -> set[str]:
    """Values considered 'auto-derived' from tap for the given non-tap op.

    A non-tap op cell is auto-derived if it matches one of these values —
    in which case the binding does not provide a distinct assignment for
    that op (just OS auto-repeat / modifier composition / uniform pass-through).
    """
    if not tap_value:
        return {''}
    if op == 'ホールド':
        return {tap_value}
    if op == 'ダブルタップ':
        return {tap_value, f'{tap_value}×2'}
    if op == 'Shift+':
        return {tap_value, f'⇧{tap_value}'}
    if op == 'Ctrl+':
        return {tap_value, f'⌃{tap_value}'}
    return set()


def _normalize_cell(value: str) -> str:
    """Convert resolve() output strings into their markdown cell form."""
    if value in ('何もしない', '&none'):
        return ''
    if value in ('フォールスルー', '&trans'):
        return '▽'
    if value.startswith('未対応'):
        return ''
    return value


def _split_cell_lines(value: str) -> list[str]:
    """Split a cell value into display lines at each ' ▸ ' separator, keeping the
    '▸' marker at the start of the continuation line. Values without a separator
    (single bindings, '▽', empty, key labels that may contain '→') are returned
    unchanged as a single-element list."""
    return value.replace(' ▸ ', '\n▸ ').split('\n')


def _disp_width(s: str) -> int:
    """Approximate display width in Excel character units, counting wide (CJK and
    other fullwidth) characters as 2 so column widths fit without auto-wrapping."""
    return sum(2 if unicodedata.east_asian_width(ch) in ('W', 'F') else 1 for ch in s)


def _compute_active_indices(layers_data: list[tuple[str, list[str]]]) -> set[int] | None:
    """Binding indices visible across the multi-layer tables: positions that are
    not `&none` in the DEFAULT layer.
    Returns None when there is no DEFAULT layer (=> show everything)."""
    default_bindings = next((b for n, b in layers_data if n == 'DEFAULT'), None)
    if default_bindings is None:
        return None
    return {i for i, b in enumerate(default_bindings) if b.strip() != '&none'}


def _build_layer_mode_table(bindings: list[str],
                            behaviors: dict, macros: dict,
                            mode: str,
                            grid, display_cols,
                            active_indices: set[int] | None = None):
    """Compute the consolidated (layer, mode) table shared by Markdown and Excel.

    Returns (header, rows):
      header: list[str] column labels ('操作', '1'.., '' for gap columns), or
              None when there is no grid/display layout.
      rows:   list of dicts in display order:
        {'kind': 'heading', 'desc': str,
         'keys': [(label, binding) | None per display column]}
        {'kind': 'detail', 'op': str,
         'values': [str per display column]}   # normalized: '', '▽', '〃', or value

    Layout rules (identical for both renderers): single column header; one heading
    row per physical row carrying the key label + binding; the タップ row always
    present and other op rows only when they differ from the auto-derived tap form;
    none/trans normalization; hidden inactive keys; skipped fully-blank rows.
    Cell strings are raw (not escaped) — the Markdown renderer escapes them.
    """
    if not grid or not display_cols:
        return None, []

    def visible_idx(cell):
        """Binding index to render for a cell, or None for a blank cell."""
        if cell is None:
            return None
        if active_indices is not None and cell not in active_indices:
            return None
        return cell

    # Header: number the key columns, leave gap columns blank.
    header = ['操作']
    num = 0
    for col in display_cols:
        if col == GAP:
            header.append('')
        else:
            num += 1
            header.append(str(num))

    non_tap_ops = ('ホールド', 'ダブルタップ', 'Shift+', 'Ctrl+')
    rows: list[dict] = []

    for i, row in enumerate(grid):
        idx_cells = [visible_idx(c) for c in row['cells']]
        if not any(c is not None for c in idx_cells):
            continue
        desc = f'Row {i + 1}'

        keys = [None if idx is None else (get_label(idx), bindings[idx])
                for idx in idx_cells]
        rows.append({'kind': 'heading', 'desc': desc, 'keys': keys})

        # Resolve each op for the visible cells, keyed by column position.
        action_by_op: dict[str, dict[int, str]] = {}
        for op in OPS:
            action_by_op[op] = {
                j: _normalize_cell(resolve(bindings[idx], behaviors, macros, op)[0])
                for j, idx in enumerate(idx_cells) if idx is not None
            }
        tap_actions = action_by_op['タップ']

        visible_ops = ['タップ']
        for op in non_tap_ops:
            cells = action_by_op[op]
            if any(cells[j] not in _auto_forms(tap_actions[j], op) for j in cells):
                visible_ops.append(op)

        for op in visible_ops:
            values: list[str] = []
            for j, idx in enumerate(idx_cells):
                if idx is None:
                    values.append('')
                    continue
                if op != 'タップ' and action_by_op[op][j] in _auto_forms(tap_actions[j], op):
                    # Derivable from the single-tap value: abbreviate (or leave
                    # blank when the tap itself is empty / does nothing).
                    values.append('' if tap_actions[j] == '' else '〃')
                    continue
                action, path = resolve(bindings[idx], behaviors, macros, op)
                values.append(_normalize_cell(action if mode == 'action' else path))
            rows.append({'kind': 'detail', 'op': op, 'values': values})

    return header, rows


HTML_ROW_BG = '#fff3cd'  # 'Row N' heading-row highlight (matches the Excel fill)

HTML_STYLE = """\
  body {
    font-family: -apple-system, "Segoe UI", "Hiragino Kaku Gothic ProN",
                 "Noto Sans JP", Meiryo, sans-serif;
    line-height: 1.6;
    color: #1f2328;
    max-width: 1400px;
    margin: 2rem auto;
    padding: 0 1.5rem;
  }
  h1 { border-bottom: 2px solid #d0d7de; padding-bottom: .3em; }
  h2 { border-bottom: 1px solid #d0d7de; padding-bottom: .3em; margin-top: 2em; }
  table {
    border-collapse: collapse;
    width: 100%;
    margin: 1em 0 2em;
    font-size: 13px;
  }
  th, td {
    border: 1px solid #d0d7de;
    padding: 4px 8px;
    text-align: center;
    vertical-align: middle;
    white-space: nowrap;   /* セル内は自動折り返しせず、明示的な改行(<br>)のみで改行する */
    font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    font-size: 12px;
  }
  thead th { background: #f6f8fa; position: sticky; top: 0; }
  /* 複数行セルは各行を左揃え、ブロック自体はセル内で中央に配置する。 */
  span.ml { display: inline-block; text-align: left; }
  /* Layer separator row inside the merged per-mode table. */
  tr.layer-row th {
    background: #e1ecf4;
    text-align: left;
    font-size: 14px;
    padding: 8px;
  }
  /* "Row N" heading rows are highlighted via inline background-color. */
  code {
    background: rgba(175,184,193,.2);
    padding: .1em .3em;
    border-radius: 4px;
    font-size: 85%;
    font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
  }
  tr[style] code { background: rgba(0,0,0,.06); }
"""


def _html_esc(s: str) -> str:
    """Escape HTML special characters."""
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _html_text(s: str) -> str:
    """Escape a table-cell text value. Backtick and backslash are rendered as
    numeric references (kept from the original Markdown-table escaping) so the
    output is stable; the leading &amp; substitution runs first so it is safe."""
    return _html_esc(s).replace('`', '&#96;').replace('\\', '&#92;')


def _html_inline(text: str) -> str:
    """Convert inline `code` spans in prose to <code>…</code>, escaping the rest."""
    holds: list[str] = []

    def code_repl(m):
        holds.append('<code>' + _html_esc(m.group(1)) + '</code>')
        return f'\x00{len(holds) - 1}\x00'

    text = re.sub(r'`([^`]*)`', code_repl, text)
    text = _html_esc(text)
    return re.sub(r'\x00(\d+)\x00', lambda m: holds[int(m.group(1))], text)


def _html_table_open(header: list[str], css_class: str | None = None) -> list[str]:
    out = [f'<table class="{css_class}">' if css_class else '<table>', '<thead>', '<tr>']
    for col in header:
        out.append(f'<th>{_html_text(col)}</th>')
    out += ['</tr>', '</thead>', '<tbody>']
    return out


def _html_table_close() -> list[str]:
    return ['</tbody>', '</table>']


def _html_layer_row(layer_name: str, n_cols: int) -> str:
    return (f'<tr class="layer-row">'
            f'<th colspan="{n_cols}">{_html_text(f"{layer_name} レイヤー")}</th>'
            f'</tr>')


def _html_body_rows(rows: list[dict]) -> list[str]:
    """Render the body of a (layer, mode) table. The 'Row N' heading rows
    carry the key label + binding and are highlighted."""
    out: list[str] = []
    for row in rows:
        if row['kind'] == 'heading':
            out.append(f'<tr style="background-color:{HTML_ROW_BG}">')
            out.append(f'<td>{_html_text(row["desc"])}</td>')
            for key in row['keys']:
                if key is None:
                    out.append('<td></td>')
                else:
                    label, binding = key
                    out.append(f'<td>{_html_text(label)}<br><code>{_html_esc(format_binding_for_display(binding))}</code></td>')
            out.append('</tr>')
        else:
            out.append('<tr>')
            out.append(f'<td>{_html_text(row["op"])}</td>')
            for value in row['values']:
                lines = _split_cell_lines(value)
                cell = '<br>'.join(_html_text(line) for line in lines)
                # 複数行セルは各行を左揃えしつつ、ブロックをセル内で中央に置く。
                if len(lines) > 1:
                    cell = f'<span class="ml">{cell}</span>'
                out.append(f'<td>{cell}</td>')
            out.append('</tr>')
    return out


def _html_table_lines(header: list[str], rows: list[dict], css_class: str | None = None) -> list[str]:
    """Render a single-layer (header + body + close) HTML table."""
    return _html_table_open(header, css_class) + _html_body_rows(rows) + _html_table_close()


def write_html(layers_data: list[tuple[str, list[str]]],
               behaviors: dict, macros: dict, output_path: Path,
               grid, display_cols) -> None:
    """Generate one standalone HTML file.
    Single layer  => H1 layer title, then H2 動作 / H2 経路.
    Multi layers  => H1 top title, H2 動作 (each layer at H3), then H2 経路."""
    body: list[str] = []

    if len(layers_data) == 1:
        layer_name, bindings = layers_data[0]
        body.append(f'<h1>{_html_inline(f"{layer_name} レイヤー キー割り当て一覧")}</h1>')
        body.append('<p>' + _html_inline(
            f'※ {len(bindings)} 個のバインディング位置を 1 表に集約。'
            f'実機の物理配列に合わせて「Row N」セクション行 + 操作行を縦に並べる（左右分割は中央の空列で分離）。'
        ) + '</p>')
        body.append('<ul>')
        body.append('<li>' + _html_inline('各 row セクション行に「キーラベル」と「バインディング (`&...`)」の 2 段表示でキー位置を示す。') + '</li>')
        body.append('<li>' + _html_inline('各表の左端 1 列が「操作」（タップ / ホールド / ダブルタップ / Shift+ / Ctrl+）または「Row N」見出し。') + '</li>')
        body.append('</ul>')
        for mode_label, mode in [('動作', 'action'), ('経路', 'path')]:
            body.append(f'<h2>{_html_inline(mode_label)}</h2>')
            header, rows = _build_layer_mode_table(bindings, behaviors, macros, mode,
                                                   grid, display_cols)
            if header is not None:
                body += _html_table_lines(header, rows)
    else:
        body.append(f'<h1>{_html_inline("キー割り当て一覧")}</h1>')
        body.append('<p>' + _html_inline(
            f'※ {len(layers_data)} 個のレイヤーのキー割り当てを 1 ファイルに集約。'
            f'各レイヤーを実機の物理配列に合わせて「動作」セクションでまとめてから「経路」セクションに進む。'
        ) + '</p>')
        body.append('<ul>')
        body.append('<li>' + _html_inline('各 row セクション行に「キーラベル」と「バインディング (`&...`)」の 2 段表示でキー位置を示す。') + '</li>')
        body.append('<li>' + _html_inline('列は物理配列の左→右順。左右分割は中央の空列で分離する。') + '</li>')
        body.append('<li>' + _html_inline('各表の左端 1 列が「操作」（タップ / ホールド / ダブルタップ / Shift+ / Ctrl+）または「Row N」見出し。') + '</li>')
        body.append('</ul>')

        # Positions that are `&none` in the DEFAULT layer are inactive and
        # hidden in every layer's table.
        active_indices = _compute_active_indices(layers_data)

        for mode_label, mode in [('動作', 'action'), ('経路', 'path')]:
            body.append(f'<h2>{_html_inline(mode_label)}</h2>')
            # Merge every layer's rows into a single table so the column widths
            # (which the browser auto-sizes per-table) line up across layers.
            shared_header: list[str] | None = None
            layer_blocks: list[tuple[str, list[dict]]] = []
            for layer_name, bindings in layers_data:
                header, rows = _build_layer_mode_table(bindings, behaviors, macros, mode,
                                                       grid, display_cols,
                                                       active_indices=active_indices)
                if header is None:
                    continue
                shared_header = header
                layer_blocks.append((layer_name, rows))
            if shared_header is None:
                continue
            body += _html_table_open(shared_header)
            for layer_name, rows in layer_blocks:
                body.append(_html_layer_row(layer_name, len(shared_header)))
                body += _html_body_rows(rows)
            body += _html_table_close()

    html = (
        '<!DOCTYPE html>\n<html lang="ja">\n<head>\n'
        '<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        '<title>キー割り当て一覧</title>\n<style>\n' + HTML_STYLE + '</style>\n</head>\n<body>\n'
        + '\n'.join(body)
        + '\n</body>\n</html>\n'
    )
    output_path.write_text(html, encoding='utf-8')


# ============================================================================
# Main
# ============================================================================

def main() -> int:
    p = argparse.ArgumentParser(
        description='Generate Excel (.xlsx) and standalone HTML (.html) docs for ZMK keymap layers '
                    '(all layers by default, or only the layer names given)'
    )
    p.add_argument('keymap', help='Path to .keymap file')
    p.add_argument('layers', nargs='*',
                   help='Zero or more layer names (e.g., VIM_NORMAL_1 VIM_NORMAL_2 VIM_VISUAL). '
                        'If omitted, all layers in the keymap are output in definition order.')
    p.add_argument('-o', '--output', help='Output .xlsx path (default: <layer>_keymap.xlsx or keymap.xlsx)')
    p.add_argument('-l', '--layout',
                   help='Path to the physical-layout JSON (default: keymap path with .json suffix)')
    args = p.parse_args()

    keymap_path = Path(args.keymap)
    if not keymap_path.is_file():
        print(f'error: keymap file not found: {keymap_path}', file=sys.stderr)
        return 1

    raw = keymap_path.read_text(encoding='utf-8')
    defines = parse_defines(raw)
    content = expand_defines(strip_comments(raw), defines)

    # Map every layer's definition-order index to its name so &lt/&mo/&to routes
    # can be rendered with the formal layer name instead of a bare number.
    all_layer_names = parse_all_layer_names(content)
    LAYER_NAMES_BY_INDEX.clear()
    LAYER_NAMES_BY_INDEX.update(dict(enumerate(all_layer_names)))

    # Default to every layer (in definition order) when none are specified.
    layer_names = args.layers if args.layers else all_layer_names
    if not layer_names:
        print('error: no layers found in keymap.', file=sys.stderr)
        return 1

    macros = parse_macros(content)
    behaviors = parse_behaviors(content)

    layers_data: list[tuple[str, list[str]]] = []
    for layer_name in layer_names:
        layer_text = parse_layer(content, layer_name)
        if layer_text is None:
            print(f'error: layer "{layer_name}" not found.', file=sys.stderr)
            return 1
        bindings = split_layer_bindings(layer_text)
        layers_data.append((layer_name, bindings))
        print(f'layer {layer_name}: {len(bindings)} bindings')

    print(f'parsed: {len(macros)} macros, {len(behaviors)} behaviors')

    # Physical layout drives the row/column arrangement to match the real board.
    total = len(layers_data[0][1]) if layers_data else 0
    layout_path = Path(args.layout) if args.layout else keymap_path.with_suffix('.json')
    coords, labels = load_physical_layout(layout_path)
    if coords and len(coords) == total:
        KEY_LABELS.clear()
        KEY_LABELS.update(labels)
        grid, display_cols = build_grid(coords)
        print(f'physical layout: {layout_path} ({len(coords)} keys)')
    else:
        grid, display_cols = legacy_grid(total)
        if coords is None:
            print(f'note: physical layout not found at {layout_path}; using keymap order',
                  file=sys.stderr)
        else:
            print(f'warning: layout key count ({len(coords)}) != bindings ({total}); '
                  'using keymap order', file=sys.stderr)

    if args.output:
        output_path = Path(args.output)
    elif len(layer_names) == 1:
        output_path = Path(f'{layer_names[0]}_keymap.xlsx')
    else:
        output_path = Path('keymap.xlsx')

    if HAVE_OPENPYXL:
        write_excel(layers_data, behaviors, macros, output_path, grid, display_cols)
        print(f'saved: {output_path}')
    else:
        print('warning: openpyxl not installed; skipping .xlsx (HTML still generated)',
              file=sys.stderr)

    html_path = output_path.with_suffix('.html')
    write_html(layers_data, behaviors, macros, html_path, grid, display_cols)
    print(f'saved: {html_path}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
