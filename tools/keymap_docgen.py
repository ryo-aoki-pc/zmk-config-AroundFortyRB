"""
keymap_docgen.py

ZMK keymap (.keymap) の指定した 1 つ以上のレイヤーの全キー割り当てを以下 2 形式で出力する
ドキュメンテーション生成ツール：

  1. Excel ファイル (.xlsx)
       - 単一レイヤー指定時: "動作" シートと "経路" シートを生成（QWERTY 物理配列）
       - 複数レイヤー指定時: 各レイヤーごとに "<layer> 動作"/"<layer> 経路" シートを生成
  2. Markdown ファイル (.md)
       - 同じ内容を Markdown の表で出力（セル内改行に <br> を使用）
       - 複数レイヤー指定時は 1 ファイル内で「## 動作」セクションに全レイヤーを並べた後、
         「## 経路」セクションに全レイヤーを再度並べる構成

それぞれの表は、キーボード物理行ごとに以下の構造を持つ：
  - 左端 1 列: 「操作」 = 単発タップ / ホールド / ダブルタップ / Shift+ / Ctrl+
  - 右側の列: その物理行のキーを QWERTY 順に並べたもの

mod-morph (LSHIFT/RSHIFT, LCTL/RCTL), tap-dance, layer-tap, momentary-layer
など標準的な ZMK behavior を解析し、再帰的に動作を解決する。

Usage:
    python keymap_docgen.py <keymap_file> <layer_name> [<layer_name> ...] [-o output.xlsx]

出力先：
    -o で指定した .xlsx と同じディレクトリ／同じベース名で .md も生成される
    （-o を省略するとレイヤー単独時は <layer>_keymap.xlsx、複数時は keymap.xlsx）

Examples:
    python tools/keymap_docgen.py config/AroundForty-RB.keymap VIM_NORMAL_1 -o KEYMAP.xlsx
    python tools/keymap_docgen.py config/AroundForty-RB.keymap VIM_NORMAL_1 VIM_NORMAL_2 VIM_VISUAL -o KEYMAP.xlsx
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:  # Excel 出力は任意。未インストールでも Markdown は生成できる
    HAVE_OPENPYXL = False


# ============================================================================
# Parsing
# ============================================================================

def strip_comments(text: str) -> str:
    text = re.sub(r'//[^\n]*', '', text)
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    return text


def find_balanced_block(content: str, start_pos: int) -> tuple[int, int] | None:
    """Find the position of the matching close brace for an open brace at start_pos."""
    if start_pos >= len(content) or content[start_pos] != '{':
        return None
    depth = 1
    i = start_pos + 1
    while i < len(content):
        if content[i] == '{':
            depth += 1
        elif content[i] == '}':
            depth -= 1
            if depth == 0:
                return (start_pos + 1, i)
        i += 1
    return None


def extract_named_block(content: str, name: str) -> str | None:
    """Extract the content of a top-level named block like 'macros { ... }'."""
    # Find 'name {'
    for m in re.finditer(r'\b' + re.escape(name) + r'\s*\{', content):
        brace_pos = m.end() - 1
        range_ = find_balanced_block(content, brace_pos)
        if range_:
            return content[range_[0]:range_[1]]
    return None


def parse_definitions(block_content: str) -> dict:
    """
    Parse 'name: label { ... };' definitions inside a block.
    Returns dict: {name: body_string}
    """
    defs = {}
    i = 0
    while i < len(block_content):
        m = re.search(r'(\w+)\s*:\s*\w+\s*\{', block_content[i:])
        if not m:
            break
        name = m.group(1)
        brace_pos = i + m.end() - 1
        range_ = find_balanced_block(block_content, brace_pos)
        if not range_:
            break
        body = block_content[range_[0]:range_[1]]
        defs[name] = body
        i = range_[1] + 1
    return defs


def parse_bindings_list(body: str) -> list[str]:
    """Extract list of bindings from a 'bindings = <...>, <...>;' style body."""
    m = re.search(r'bindings\s*=\s*([^;]+);', body, re.DOTALL)
    if not m:
        return []
    text = m.group(1)
    return [s.strip() for s in re.findall(r'<([^>]+)>', text)]


def parse_compatible(body: str) -> str | None:
    m = re.search(r'compatible\s*=\s*"([^"]+)"', body)
    return m.group(1) if m else None


def parse_mods(body: str) -> str | None:
    m = re.search(r'mods\s*=\s*<\(?([^>)]+)\)?>', body)
    return m.group(1).strip() if m else None


def parse_macros(content: str) -> dict:
    block = extract_named_block(content, 'macros')
    if not block:
        return {}
    defs = parse_definitions(block)
    return {name: {'bindings': parse_bindings_list(body)} for name, body in defs.items()}


def parse_behaviors(content: str) -> dict:
    block = extract_named_block(content, 'behaviors')
    if not block:
        return {}
    defs = parse_definitions(block)
    out = {}
    for name, body in defs.items():
        out[name] = {
            'compatible': parse_compatible(body),
            'bindings': parse_bindings_list(body),
            'mods': parse_mods(body),
        }
    return out


def parse_layer(content: str, layer_name: str) -> str | None:
    """Extract the raw bindings text of a named layer."""
    keymap_block = extract_named_block(content, 'keymap')
    if not keymap_block:
        return None
    defs = {}
    # layers are 'NAME { ... }' (no label after colon)
    i = 0
    while i < len(keymap_block):
        m = re.search(r'(\w+)\s*\{', keymap_block[i:])
        if not m:
            break
        name = m.group(1)
        brace_pos = i + m.end() - 1
        range_ = find_balanced_block(keymap_block, brace_pos)
        if not range_:
            break
        body = keymap_block[range_[0]:range_[1]]
        if name == layer_name:
            m2 = re.search(r'bindings\s*=\s*<(.+?)>', body, re.DOTALL)
            if m2:
                return m2.group(1)
        i = range_[1] + 1
    return None


def split_layer_bindings(text: str) -> list[str]:
    """Split a layer bindings string into individual binding strings (each starts with '&')."""
    text = re.sub(r'\s+', ' ', text).strip()
    out = []
    i = 0
    n = len(text)
    while i < n:
        if text[i] == '&':
            j = i + 1
            while j < n and text[j] != '&':
                j += 1
            out.append(text[i:j].strip())
            i = j
        else:
            i += 1
    return out


# ============================================================================
# Keycode formatting
# ============================================================================

KEYCODE_LABELS = {
    'LEFT': '←', 'RIGHT': '→', 'UP_ARROW': '↑', 'UP': '↑', 'DOWN': '↓',
    'HOME': 'HOME', 'END': 'END', 'ENTER': 'ENTER', 'DELETE': 'DEL',
    'BACKSPACE': 'BS', 'TAB': 'TAB', 'SPACE': 'SPACE', 'ESCAPE': 'ESC',
    'PAGE_UP': 'PgUp', 'PAGE_DOWN': 'PgDn',
    'LCTRL': 'LCtrl', 'RCTRL': 'RCtrl',
    'LSHIFT': 'LShift', 'RSHIFT': 'RShift',
    'LEFT_SHIFT': 'LShift', 'RIGHT_SHIFT': 'RShift',
    'LEFT_CONTROL': 'LCtrl', 'RIGHT_CONTROL': 'RCtrl',
    'LEFT_ALT': 'LAlt', 'RIGHT_ALT': 'RAlt',
    'LEFT_WIN': 'LWin', 'RIGHT_WIN': 'RWin',
    'GREATER_THAN': '>', 'LESS_THAN': '<',
    'COMMA': ',', 'PERIOD': '.', 'SEMICOLON': ';', 'SLASH': '/',
    'SINGLE_QUOTE': "'", 'GRAVE': '`', 'BACKSLASH': '\\', 'EQUAL': '=',
    'MINUS': '-', 'LEFT_BRACKET': '[', 'RIGHT_BRACKET': ']',
    'N0': '0', 'N1': '1', 'N2': '2', 'N3': '3', 'N4': '4',
    'N5': '5', 'N6': '6', 'N7': '7', 'N8': '8', 'N9': '9',
}

MOD_PREFIX = {
    'LC': 'Ctrl', 'RC': 'Ctrl',
    'LS': 'Shift', 'RS': 'Shift',
    'LA': 'Alt', 'RA': 'Alt',
    'LG': 'Win', 'RG': 'Win',
}


def format_keycode(kc: str) -> str:
    kc = kc.strip()
    m = re.match(r'(LC|LS|LA|LG|RC|RS|RA|RG)\((.+)\)$', kc)
    if m:
        return f"{MOD_PREFIX[m.group(1)]}+{format_keycode(m.group(2))}"
    return KEYCODE_LABELS.get(kc, kc)


# ============================================================================
# Resolution: binding × operation -> (action_description, path)
# ============================================================================

OPS = ('単発タップ', 'ホールド', 'ダブルタップ', 'Shift+', 'Ctrl+')


def resolve(binding: str, behaviors: dict, macros: dict, op: str, depth: int = 0) -> tuple[str, str]:
    """Resolve a binding string for the given operation."""
    if depth > 10:
        return ('再帰深度超過', binding)
    b = binding.strip()

    if b == '&none':
        return ('何もしない', '&none')

    if b == '&trans':
        return ('フォールスルー', '&trans')

    # &kp X
    m = re.match(r'&kp\s+(.+)$', b)
    if m:
        kc = m.group(1).strip()
        label = format_keycode(kc)
        path = f'&kp {kc}'
        if op == '単発タップ':
            return (label, path)
        if op == 'ホールド':
            return (label, path)
        if op == 'ダブルタップ':
            return (f'{label}×2', path)
        if op == 'Shift+':
            return (f'⇧{label}', path)
        if op == 'Ctrl+':
            return (f'⌃{label}', path)

    # &mt MOD KEY
    m = re.match(r'&mt\s+(\S+)\s+(.+)$', b)
    if m:
        mod, key = m.group(1).strip(), m.group(2).strip()
        key_label = format_keycode(key)
        mod_label = format_keycode(mod)
        path = f'&mt {mod} {key}'
        if op == '単発タップ':
            return (key_label, path)
        if op == 'ホールド':
            return (mod_label, path)
        if op == 'ダブルタップ':
            return (f'{key_label}×2', path)
        if op == 'Shift+':
            return (f'⇧{key_label}', path)
        if op == 'Ctrl+':
            return (f'⌃{key_label}', path)

    # &lt LAYER KEY
    m = re.match(r'&lt\s+(\d+)\s+(.+)$', b)
    if m:
        layer, key = m.group(1), m.group(2).strip()
        key_label = format_keycode(key)
        path = f'&lt {layer} {key}'
        if op == '単発タップ':
            return (key_label, path)
        if op == 'ホールド':
            return (f'L{layer}', path)
        if op == 'ダブルタップ':
            return (f'{key_label}×2', path)
        if op == 'Shift+':
            return (f'⇧{key_label}', path)
        if op == 'Ctrl+':
            return (f'⌃{key_label}', path)

    # &mo X
    m = re.match(r'&mo\s+(\d+)$', b)
    if m:
        layer = m.group(1)
        return (f'L{layer}', f'&mo {layer}')

    # &to X
    m = re.match(r'&to\s+(\d+)$', b)
    if m:
        layer = m.group(1)
        return (f'⇒L{layer}', f'&to {layer}')

    # Custom behavior / macro reference like &mm_vim_g, &td_vim_d, &macro_vim_dd
    if b.startswith('&'):
        name = b[1:].split()[0]
        if name in behaviors:
            return resolve_behavior(name, behaviors, macros, op, depth + 1)
        if name in macros:
            return resolve_macro(name, behaviors, macros, op, depth + 1)

    return (f'未対応: {b}', b)


def resolve_behavior(name: str, behaviors: dict, macros: dict, op: str, depth: int) -> tuple[str, str]:
    beh = behaviors[name]
    compat = beh['compatible']
    bindings = beh['bindings']
    mods = beh['mods'] or ''

    if compat == 'zmk,behavior-mod-morph':
        is_shift = 'LSFT' in mods or 'RSFT' in mods
        is_ctrl = 'LCTL' in mods or 'RCTL' in mods

        if op in ('単発タップ', 'ダブルタップ', 'ホールド'):
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, op, depth)
            return (sub_a, f'{name}[0] → {sub_p}')

        if op == 'Shift+':
            if is_shift:
                sub_a, sub_p = resolve(bindings[1], behaviors, macros, '単発タップ', depth)
                return (sub_a, f'{name}[1] → {sub_p}')
            else:
                sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Shift+', depth)
                return (sub_a, f'{name}[0] → {sub_p}')

        if op == 'Ctrl+':
            if is_ctrl:
                sub_a, sub_p = resolve(bindings[1], behaviors, macros, '単発タップ', depth)
                return (sub_a, f'{name}[1] → {sub_p}')
            else:
                sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Ctrl+', depth)
                return (sub_a, f'{name}[0] → {sub_p}')

    if compat == 'zmk,behavior-tap-dance':
        if op == '単発タップ':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, '単発タップ', depth)
            return (sub_a, f'{name}[0] → {sub_p}')
        if op == 'ホールド':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'ホールド', depth)
            return (sub_a, f'{name}[0] → {sub_p}')
        if op == 'ダブルタップ':
            sub_a, sub_p = resolve(bindings[1], behaviors, macros, '単発タップ', depth)
            return (sub_a, f'{name}[1] → {sub_p}')
        if op == 'Shift+':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Shift+', depth)
            return (sub_a, f'{name}[0] → {sub_p}')
        if op == 'Ctrl+':
            sub_a, sub_p = resolve(bindings[0], behaviors, macros, 'Ctrl+', depth)
            return (sub_a, f'{name}[0] → {sub_p}')

    return (f'未対応 behavior: {compat}', name)


def resolve_macro(name: str, behaviors: dict, macros: dict, op: str, depth: int) -> tuple[str, str]:
    summary = summarize_macro(macros[name]['bindings'])
    if op == 'ダブルタップ':
        return (f'{summary}×2', name)
    if op == 'Shift+':
        return (f'⇧{summary}', name)
    if op == 'Ctrl+':
        return (f'⌃{summary}', name)
    return (summary, name)


def summarize_macro(bindings: list[str]) -> str:
    parts = []
    for b in bindings:
        b = b.strip()
        if b.startswith('&macro_wait_time'):
            continue
        if b.startswith('&kp '):
            for kc in re.findall(r'&kp\s+([A-Z0-9_()]+)', b):
                parts.append(format_keycode(kc))
        elif b.startswith('&macro_release'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(f'{format_keycode(inner.group(1))} 解放')
        elif b.startswith('&macro_press'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(f'{format_keycode(inner.group(1))} 押下')
        elif b.startswith('&macro_tap'):
            inner = re.search(r'&kp\s+(\S+)', b)
            if inner:
                parts.append(format_keycode(inner.group(1)))
        elif b.startswith('&to '):
            parts.append(f'レイヤー {b[4:].strip()} へ')
        else:
            parts.append(b)
    return ' ▸ '.join(parts)


# ============================================================================
# Visual label (per physical key index, not per column position)
# ============================================================================

# Label for each binding index 0..41 = the DEFAULT-layer physical key identity.
# Keyed by index (not by column slot) so reordering columns into physical order
# can never desync a label from its binding.
KEY_LABELS = {
    0: 'Q', 1: 'W', 2: 'E', 3: 'R', 4: 'T', 5: 'Y', 6: 'U', 7: 'I', 8: 'O', 9: 'P',
    10: 'A', 11: 'S', 12: 'D', 13: 'F', 14: 'G', 15: 'H', 16: 'J', 17: 'K', 18: 'L', 19: '-',
    20: 'Z', 21: 'X', 22: 'C', 23: 'V', 24: 'B', 25: '(center mo7)',
    26: 'N', 27: 'M', 28: ',', 29: '.', 30: '/',
    31: 'mo6 (L outer)', 32: 'LWin', 33: 'LAlt', 34: 'SPACE', 35: 'SPACE',
    36: 'mo1 (L center)', 37: 'lt1 ENTER', 38: 'mo7 (raised)', 39: 'mo2',
    40: 'mo6 (R)', 41: 'mo6 (R outer)',
}


def get_label(idx: int) -> str:
    return KEY_LABELS.get(idx, f'pos {idx}')


# ============================================================================
# Excel output
# ============================================================================

ROW_DESCRIPTIONS = {
    1: 'Row 1 (QWERTY 上段)',
    2: 'Row 2 (home row)',
    3: 'Row 3 (Z row)',
    4: 'Row 4 (thumb)',
}


def get_row_layout(total: int) -> list[tuple[int, int]]:
    """Determine the physical row layout based on total binding count."""
    if total == 42:
        return [(1, 10), (2, 10), (3, 11), (4, 11)]
    return [(0, total)]  # fallback: linear


# ============================================================================
# Physical layout grid (drives row/column arrangement from real coordinates)
# ============================================================================

GAP = 'GAP'  # sentinel marking a blank split-gap display column


def load_physical_layout(path: Path) -> list[tuple[float, float]] | None:
    """Parse a ZMK physical-layout JSON into an ordered (x, y) list per key index.

    The JSON's `layout` array is in the same order as the keymap bindings, so
    entry i is the physical position of binding i. Returns None when the file is
    missing or cannot be parsed (caller falls back to keymap-order rendering).
    """
    try:
        data = json.loads(Path(path).read_text(encoding='utf-8'))
    except (OSError, ValueError):
        return None
    layouts = data.get('layouts') if isinstance(data, dict) else None
    if not layouts:
        return None
    layout = layouts.get('default_layout') or next(iter(layouts.values()))
    entries = layout.get('layout') if isinstance(layout, dict) else None
    if not entries:
        return None
    coords: list[tuple[float, float]] = []
    for e in entries:
        try:
            coords.append((float(e['x']), float(e['y'])))
        except (KeyError, TypeError, ValueError):
            return None
    return coords or None


def build_grid(coords: list[tuple[float, float]]):
    """Build physical rows + display columns from per-index (x, y) coordinates.

    Rows are grouped by distinct y (top to bottom); columns are the distinct x
    values (left to right) with a single blank GAP column inserted at the widest
    horizontal gap (the split between the two halves). Returns
    (rows, display_cols) where:
      - display_cols: left-to-right list of x-values (float) and GAP sentinels.
      - rows: list of {'y': y, 'cells': [idx | None per display column]} where a
        None cell is an empty position, the split gap, or a missing key.
    """
    if not coords:
        return None, None
    xs = sorted({x for x, _ in coords})
    ys = sorted({y for _, y in coords})

    # Insert a gap column at the largest consecutive x-gap, if one clearly exists.
    diffs = [xs[i + 1] - xs[i] for i in range(len(xs) - 1)]
    if diffs and max(diffs) > min(diffs):
        gap_after = diffs.index(max(diffs))
    else:
        gap_after = -1

    display_cols: list = []
    for i, x in enumerate(xs):
        display_cols.append(x)
        if i == gap_after:
            display_cols.append(GAP)

    pos = {(x, y): idx for idx, (x, y) in enumerate(coords)}
    rows = []
    for y in ys:
        cells = [None if col == GAP else pos.get((col, y)) for col in display_cols]
        rows.append({'y': y, 'cells': cells})
    return rows, display_cols


def legacy_grid(total: int):
    """Fallback grid (no physical coords): sequential keymap-order chunking."""
    layout = get_row_layout(total)
    max_w = max((count for _, count in layout), default=0)
    rows = []
    idx = 0
    for _, count in layout:
        cells = [idx + p for p in range(count)] + [None] * (max_w - count)
        rows.append({'y': None, 'cells': cells})
        idx += count
    return rows, list(range(max_w))


def _write_mode_sheet(ws, layers_data: list[tuple[str, list[str]]],
                      behaviors: dict, macros: dict, mode: str,
                      grid, display_cols, active_indices, is_single: bool) -> None:
    """Write one mode sheet ('動作' or '経路') laid out to match the HTML/Markdown
    tables: a single column header per table, '■ Row N' heading rows (highlighted)
    that carry the key label + binding, then only the relevant operation rows.
    Multiple layers are stacked vertically with a layer-name heading before each.
    """
    title_font = Font(bold=True, size=14, name='Yu Gothic UI')
    layer_font = Font(bold=True, size=12, name='Yu Gothic UI')
    header_font = Font(bold=True, size=10, name='Yu Gothic UI')
    header_fill = PatternFill('solid', start_color='F6F8FA', end_color='F6F8FA', fill_type='solid')
    row_font = Font(bold=True, size=10, name='Yu Gothic UI')
    row_fill = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    op_font = Font(bold=True, size=9, name='Yu Gothic UI')
    body_font = Font(size=9, name='Yu Gothic UI')

    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    mode_label = '動作' if mode == 'action' else '経路'
    n_cols = len(display_cols)

    # Column widths
    ws.column_dimensions['A'].width = 22
    col_width = 18 if mode == 'action' else 30
    for j in range(n_cols):
        width = 3 if display_cols[j] == GAP else col_width
        ws.column_dimensions[get_column_letter(2 + j)].width = width

    # Sheet title
    if is_single:
        layer_name = layers_data[0][0]
        title = f'{layer_name} レイヤー - {mode_label}'
    else:
        title = f'キー割り当て一覧 - {mode_label}'
    c = ws.cell(1, 1, title)
    c.font = title_font
    r = 3

    for layer_name, bindings in layers_data:
        if not is_single:
            c = ws.cell(r, 1, f'{layer_name} レイヤー')
            c.font = layer_font
            r += 1

        header, rows = _build_layer_mode_table(bindings, behaviors, macros, mode,
                                               grid, display_cols, active_indices)
        if header is None:
            r += 1
            continue

        # Column header row (操作 | 1 | 2 | ...).
        for col, text in enumerate(header, start=1):
            c = ws.cell(r, col, text)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
        r += 1

        for row in rows:
            if row['kind'] == 'heading':
                c = ws.cell(r, 1, f'■ {row["desc"]}')
                c.font = row_font
                c.fill = row_fill
                c.alignment = Alignment(horizontal='left', vertical='center')
                c.border = border
                for j, key in enumerate(row['keys']):
                    value = '' if key is None else f'{key[0]}\n{key[1]}'
                    cc = ws.cell(r, 2 + j, value)
                    cc.font = body_font
                    cc.fill = row_fill
                    cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cc.border = border
                ws.row_dimensions[r].height = 30
            else:
                c = ws.cell(r, 1, row['op'])
                c.font = op_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border
                for j, value in enumerate(row['values']):
                    cc = ws.cell(r, 2 + j, value)
                    cc.font = body_font
                    cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cc.border = border
            r += 1

        r += 1  # blank row between stacked tables


def write_excel(layers_data: list[tuple[str, list[str]]],
                behaviors: dict, macros: dict, output_path: Path,
                grid, display_cols) -> None:
    """Generate one Excel file matching the HTML/Markdown layout: two sheets
    '動作' / '経路', each stacking every layer's table vertically."""
    wb = Workbook()
    is_single = len(layers_data) == 1
    active_indices = None if is_single else _compute_active_indices(layers_data)
    first = True
    for _, mode in [('動作', 'action'), ('経路', 'path')]:
        sheet_name = '動作' if mode == 'action' else '経路'
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        _write_mode_sheet(ws, layers_data, behaviors, macros, mode,
                          grid, display_cols, active_indices, is_single)
    wb.save(output_path)


# ============================================================================
# Markdown output
# ============================================================================

def _escape_md_cell(s) -> str:
    """Escape a value so it can safely appear inside a Markdown table cell."""
    if s is None:
        return ''
    return (str(s)
            .replace('\\', '&#92;')
            .replace('|', '\\|')
            .replace('`', '&#96;')
            .replace('\n', '<br>'))


def _auto_forms(tap_value: str, op: str) -> set[str]:
    """Values considered 'auto-derived' from tap for the given non-tap op.

    A non-tap op cell is auto-derived if it matches one of these values —
    in which case the binding does not provide a distinct assignment for
    that op (just OS auto-repeat / modifier composition / uniform pass-through).
    """
    if not tap_value:
        return {''}
    if op == 'ホールド':
        return {tap_value}
    if op == 'ダブルタップ':
        return {tap_value, f'{tap_value}×2'}
    if op == 'Shift+':
        return {tap_value, f'⇧{tap_value}'}
    if op == 'Ctrl+':
        return {tap_value, f'⌃{tap_value}'}
    return set()


def _normalize_cell(value: str) -> str:
    """Convert resolve() output strings into their markdown cell form."""
    if value in ('何もしない', '&none'):
        return ''
    if value in ('フォールスルー', '&trans'):
        return '▽'
    if value.startswith('未対応'):
        return ''
    return value


def _compute_active_indices(layers_data: list[tuple[str, list[str]]]) -> set[int] | None:
    """Binding indices visible across the multi-layer tables: positions that are
    not `&none` in the DEFAULT layer and not the `&mo 7` center/raised thumb keys.
    Returns None when there is no DEFAULT layer (=> show everything)."""
    default_bindings = next((b for n, b in layers_data if n == 'DEFAULT'), None)
    if default_bindings is None:
        return None
    always_hidden = {i for i, b in enumerate(default_bindings) if b.strip() == '&mo 7'}
    return {i for i, b in enumerate(default_bindings)
            if b.strip() != '&none' and i not in always_hidden}


def _build_layer_mode_table(bindings: list[str],
                            behaviors: dict, macros: dict,
                            mode: str,
                            grid, display_cols,
                            active_indices: set[int] | None = None):
    """Compute the consolidated (layer, mode) table shared by Markdown and Excel.

    Returns (header, rows):
      header: list[str] column labels ('操作', '1'.., '' for gap columns), or
              None when there is no grid/display layout.
      rows:   list of dicts in display order:
        {'kind': 'heading', 'desc': str,
         'keys': [(label, binding) | None per display column]}
        {'kind': 'detail', 'op': str,
         'values': [str per display column]}   # normalized: '', '▽', '〃', or value

    Layout rules (identical for both renderers): single column header; one heading
    row per physical row carrying the key label + binding; the 単発タップ row always
    present and other op rows only when they differ from the auto-derived tap form;
    none/trans normalization; hidden inactive keys; skipped fully-blank rows.
    Cell strings are raw (not escaped) — the Markdown renderer escapes them.
    """
    if not grid or not display_cols:
        return None, []

    def visible_idx(cell):
        """Binding index to render for a cell, or None for a blank cell."""
        if cell is None:
            return None
        if active_indices is not None and cell not in active_indices:
            return None
        return cell

    # Header: number the key columns, leave gap columns blank.
    header = ['操作']
    num = 0
    for col in display_cols:
        if col == GAP:
            header.append('')
        else:
            num += 1
            header.append(str(num))

    non_tap_ops = ('ホールド', 'ダブルタップ', 'Shift+', 'Ctrl+')
    rows: list[dict] = []

    for i, row in enumerate(grid):
        idx_cells = [visible_idx(c) for c in row['cells']]
        if not any(c is not None for c in idx_cells):
            continue
        desc = ROW_DESCRIPTIONS.get(i + 1, f'Row {i + 1}')

        keys = [None if idx is None else (get_label(idx), bindings[idx])
                for idx in idx_cells]
        rows.append({'kind': 'heading', 'desc': desc, 'keys': keys})

        # Resolve each op for the visible cells, keyed by column position.
        action_by_op: dict[str, dict[int, str]] = {}
        for op in OPS:
            action_by_op[op] = {
                j: _normalize_cell(resolve(bindings[idx], behaviors, macros, op)[0])
                for j, idx in enumerate(idx_cells) if idx is not None
            }
        tap_actions = action_by_op['単発タップ']

        visible_ops = ['単発タップ']
        for op in non_tap_ops:
            cells = action_by_op[op]
            if any(cells[j] not in _auto_forms(tap_actions[j], op) for j in cells):
                visible_ops.append(op)

        for op in visible_ops:
            values: list[str] = []
            for j, idx in enumerate(idx_cells):
                if idx is None:
                    values.append('')
                    continue
                if op != '単発タップ' and action_by_op[op][j] in _auto_forms(tap_actions[j], op):
                    # Derivable from the single-tap value: abbreviate (or leave
                    # blank when the tap itself is empty / does nothing).
                    values.append('' if tap_actions[j] == '' else '〃')
                    continue
                action, path = resolve(bindings[idx], behaviors, macros, op)
                values.append(_normalize_cell(action if mode == 'action' else path))
            rows.append({'kind': 'detail', 'op': op, 'values': values})

    return header, rows


def _markdown_layer_mode_rows(bindings: list[str],
                              behaviors: dict, macros: dict,
                              mode: str,
                              grid, display_cols,
                              active_indices: set[int] | None = None) -> list[str]:
    """Render the shared (layer, mode) table (see `_build_layer_mode_table`) as
    Markdown table lines, with a trailing blank line."""
    header, rows = _build_layer_mode_table(bindings, behaviors, macros, mode,
                                           grid, display_cols, active_indices)
    if header is None:
        return []

    lines: list[str] = []
    lines.append('| ' + ' | '.join(header) + ' |')
    lines.append('|' + '|'.join(['---'] * len(header)) + '|')

    for row in rows:
        if row['kind'] == 'heading':
            cells = [f'■ {_escape_md_cell(row["desc"])}']
            for key in row['keys']:
                if key is None:
                    cells.append('')
                else:
                    label, binding = key
                    cells.append(f'{_escape_md_cell(label)}<br>`{_escape_md_cell(binding)}`')
            lines.append('| ' + ' | '.join(cells) + ' |')
        else:
            cells = [_escape_md_cell(row['op'])]
            for value in row['values']:
                cells.append(value if value in ('', '▽', '〃') else _escape_md_cell(value))
            lines.append('| ' + ' | '.join(cells) + ' |')

    lines.append('')
    return lines


def write_markdown(layers_data: list[tuple[str, list[str]]],
                   behaviors: dict, macros: dict, output_path: Path,
                   grid, display_cols) -> None:
    """Generate one Markdown file.
    Single layer  => H1 layer title, then H2 動作 / H2 経路.
    Multi layers  => H1 top title, H2 動作 (each layer at H3), then H2 経路 (each layer at H3)."""
    lines: list[str] = []

    if len(layers_data) == 1:
        layer_name, bindings = layers_data[0]
        lines.append(f'# {layer_name} レイヤー キー割り当て一覧')
        lines.append('')
        lines.append(
            f'※ {len(bindings)} 個のバインディング位置を 1 表に集約。'
            f'実機の物理配列に合わせて「■ Row N」セクション行 + 操作行を縦に並べる（左右分割は中央の空列で分離）。'
        )
        lines.append('')
        lines.append('- 各 row セクション行に「キーラベル」と「バインディング (`&...`)」の 2 段表示でキー位置を示す。')
        lines.append('- 各表の左端 1 列が「操作」（単発タップ / ホールド / ダブルタップ / Shift+ / Ctrl+）または「■ Row N」見出し。')
        lines.append('')
        for mode_label, mode in [('動作', 'action'), ('経路', 'path')]:
            lines.append(f'## {mode_label}')
            lines.append('')
            lines.extend(_markdown_layer_mode_rows(bindings, behaviors, macros, mode,
                                                   grid, display_cols))
    else:
        lines.append('# キー割り当て一覧')
        lines.append('')
        lines.append(
            f'※ {len(layers_data)} 個のレイヤーのキー割り当てを 1 ファイルに集約。'
            f'各レイヤーを実機の物理配列に合わせて「動作」セクションでまとめてから「経路」セクションに進む。'
        )
        lines.append('')
        lines.append('- 各 row セクション行に「キーラベル」と「バインディング (`&...`)」の 2 段表示でキー位置を示す。')
        lines.append('- 列は物理配列の左→右順。左右分割は中央の空列で分離する。')
        lines.append('- 各表の左端 1 列が「操作」（単発タップ / ホールド / ダブルタップ / Shift+ / Ctrl+）または「■ Row N」見出し。')
        lines.append('')

        # Positions that are `&none` in DEFAULT, plus the `&mo 7` center-column
        # and raised thumb keys, are hidden in every layer's table.
        active_indices = _compute_active_indices(layers_data)

        for mode_label, mode in [('動作', 'action'), ('経路', 'path')]:
            lines.append(f'## {mode_label}')
            lines.append('')
            for layer_name, bindings in layers_data:
                lines.append(f'### {layer_name} レイヤー')
                lines.append('')
                lines.extend(_markdown_layer_mode_rows(bindings, behaviors, macros, mode,
                                                       grid, display_cols,
                                                       active_indices=active_indices))

    output_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


# ============================================================================
# Main
# ============================================================================

def main() -> int:
    p = argparse.ArgumentParser(
        description='Generate Excel (.xlsx) and Markdown (.md) docs for one or more ZMK keymap layers'
    )
    p.add_argument('keymap', help='Path to .keymap file')
    p.add_argument('layers', nargs='+',
                   help='One or more layer names (e.g., VIM_NORMAL_1 VIM_NORMAL_2 VIM_VISUAL)')
    p.add_argument('-o', '--output', help='Output .xlsx path (default: <layer>_keymap.xlsx or keymap.xlsx)')
    p.add_argument('-l', '--layout',
                   help='Path to the physical-layout JSON (default: keymap path with .json suffix)')
    args = p.parse_args()

    keymap_path = Path(args.keymap)
    if not keymap_path.is_file():
        print(f'error: keymap file not found: {keymap_path}', file=sys.stderr)
        return 1

    content = strip_comments(keymap_path.read_text(encoding='utf-8'))

    macros = parse_macros(content)
    behaviors = parse_behaviors(content)

    layers_data: list[tuple[str, list[str]]] = []
    for layer_name in args.layers:
        layer_text = parse_layer(content, layer_name)
        if layer_text is None:
            print(f'error: layer "{layer_name}" not found.', file=sys.stderr)
            return 1
        bindings = split_layer_bindings(layer_text)
        layers_data.append((layer_name, bindings))
        print(f'layer {layer_name}: {len(bindings)} bindings')

    print(f'parsed: {len(macros)} macros, {len(behaviors)} behaviors')

    # Physical layout drives the row/column arrangement to match the real board.
    total = len(layers_data[0][1]) if layers_data else 0
    layout_path = Path(args.layout) if args.layout else keymap_path.with_suffix('.json')
    coords = load_physical_layout(layout_path)
    if coords and len(coords) == total:
        grid, display_cols = build_grid(coords)
        print(f'physical layout: {layout_path} ({len(coords)} keys)')
    else:
        grid, display_cols = legacy_grid(total)
        if coords is None:
            print(f'note: physical layout not found at {layout_path}; using keymap order',
                  file=sys.stderr)
        else:
            print(f'warning: layout key count ({len(coords)}) != bindings ({total}); '
                  'using keymap order', file=sys.stderr)

    if args.output:
        output_path = Path(args.output)
    elif len(args.layers) == 1:
        output_path = Path(f'{args.layers[0]}_keymap.xlsx')
    else:
        output_path = Path('keymap.xlsx')

    if HAVE_OPENPYXL:
        write_excel(layers_data, behaviors, macros, output_path, grid, display_cols)
        print(f'saved: {output_path}')
    else:
        print('warning: openpyxl not installed; skipping .xlsx (Markdown still generated)',
              file=sys.stderr)

    md_path = output_path.with_suffix('.md')
    write_markdown(layers_data, behaviors, macros, md_path, grid, display_cols)
    print(f'saved: {md_path}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
